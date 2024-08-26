from flask import Flask, flash, render_template, request, redirect, url_for, send_file, jsonify
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime, date
import os, glob, time
from dateutil import parser

app = Flask(__name__)
app.secret_key = 'key_DQApp'

PROCESSED_FOLDER = 'processed'
DATA_QUALITY_FOLDER = 'data quality'
INDIVIDUAL_BORROWER_FOLDER = 'individual_borrower'
CORPORATE_BORROWER_FOLDER = 'corporate_borrower'
CREDIT_INFORMATION_FOLDER = 'credit_information'
GUARANTORS_INFORMATION_FOLDER = 'guarantors_information'
PRINCIPAL_OFFICERS_FOLDER = 'principal_officers'

# Ensure the upload folder exists
os.makedirs(PROCESSED_FOLDER, exist_ok=True)
os.makedirs(DATA_QUALITY_FOLDER, exist_ok=True)
os.makedirs(INDIVIDUAL_BORROWER_FOLDER, exist_ok=True)
os.makedirs(CORPORATE_BORROWER_FOLDER, exist_ok=True)
os.makedirs(CREDIT_INFORMATION_FOLDER, exist_ok=True)
os.makedirs(GUARANTORS_INFORMATION_FOLDER, exist_ok=True)
os.makedirs(PRINCIPAL_OFFICERS_FOLDER, exist_ok=True)

app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER
app.config['DATA_QUALITY_FOLDER'] = DATA_QUALITY_FOLDER
app.config['INDIVIDUAL_BORROWER_FOLDER'] = INDIVIDUAL_BORROWER_FOLDER
app.config['CORPORATE_BORROWER_FOLDER'] = CORPORATE_BORROWER_FOLDER
app.config['CREDIT_INFORMATION_FOLDER'] = CREDIT_INFORMATION_FOLDER
app.config['GUARANTORS_INFORMATION_FOLDER'] = GUARANTORS_INFORMATION_FOLDER
app.config['PRINCIPAL_OFFICERS_FOLDER'] = PRINCIPAL_OFFICERS_FOLDER

app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

# Define the specific names to check for in the file name
SPECIFIC_NAME = [
    "Individual-Borrower", "Credit-Information", "Corporate-Borrower", 
    "Guarantors-Information", "Principal-Officers"
]

# Function to check allowed file
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Function to check if the specific name is part of the file name
def allowed_file_name(filename):
    return any(name in filename for name in SPECIFIC_NAME)

# Function to process the uploaded file based on its type
def process_uploaded_file(file_path):
    if "Individual-Borrower" in file_path:
        process_individual_borrower(file_path)
    elif "Credit-Information" in file_path:
        process_credit_information(file_path)
    elif "Corporate-Borrower" in file_path:
        process_corporate_borrower(file_path)
    elif "Principal-Officers" in file_path:
        process_principal_officers(file_path)
    elif "Guarantors-Information" in file_path:
        process_guarantors_information(file_path)
    else:
        flash("Unknown file type.", 'danger')

# Function to parse and format date columns
def format_date_columns(sheet, date_columns):
    for col_letter in date_columns:
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f'{col_letter}{row}']
            if isinstance(cell.value, str):
                try:
                    parsed_date = parser.parse(cell.value, dayfirst=True)
                    formatted_date = parsed_date.strftime('%d-%b-%Y')
                    cell.value = formatted_date
                except (ValueError, TypeError):
                    continue
            elif isinstance(cell.value, (datetime, date)):
                cell.value = cell.value.strftime('%d-%b-%Y')

# Function to remove duplicates
def remove_duplicates(sheet):
    seen_rows = set()
    rows_to_delete = set()

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_key = tuple(row)
        if row_key in seen_rows:
            rows_to_delete.add(row_idx)
        else:
            seen_rows.add(row_key)

    if rows_to_delete:
        for row in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row)
        return True
    else:
        return False
    

# Function to clean cell values
def clean_cell_value(value):
    changes = {
        'exponential': 0,
        'pipe': 0,
        'carriage_return': 0,
        'line_break': 0
    }

    if isinstance(value, str):
        if 'E+' in value:
            value = value.replace('E+', '')
            changes['exponential'] += 1
        if '|' in value:
            value = value.replace('|', '')
            changes['pipe'] += 1
        if '\r' in value:
            value = value.replace('\r', '')
            changes['carriage_return'] += 1
        if '\n' in value:
            value = value.replace('\n', '')
            changes['line_break'] += 1

    return value, changes

# Function to clean the entire worksheet
def clean_worksheet(sheet):
    total_changes = {
        'exponential': 0,
        'pipe': 0,
        'carriage_return': 0,
        'line_break': 0
    }

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                cleaned_value, changes = clean_cell_value(cell.value)
                cell.value = cleaned_value
                for key in total_changes:
                    total_changes[key] += changes[key]

    return total_changes

# Function to print cleaning summary
def print_summary(changes):
    if any(changes.values()):
        flash("Cleaning Summary:",'success')
        if changes['exponential'] > 0:
            flash(f"Removed {changes['exponential']} exponentials (E+).",'success')
        if changes['pipe'] > 0:
            flash(f"Removed {changes['pipe']} pipe symbols (|).",'success')
        if changes['carriage_return'] > 0:
            flash(f"Removed {changes['carriage_return']} carriage returns (\\r).",'success')
        if changes['line_break'] > 0:
            flash(f"Removed {changes['line_break']} line breaks (\\n).",'success')
    else:
        flash(f"No cleaning needed.",'success')

# Function to check for and handle blank rows in specified columns
def check_blank_rows_in_columns(sheet, column_letters):
    blank_rows = []
    blank_row_indices = []

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        is_blank = False
        row_num = row[0].row
        row_values = [cell.value for cell in row]
        for col_letter in column_letters:
            col_idx = column_index_from_string(col_letter) - 1
            cell_value = row[col_idx].value
            if cell_value in [None, '']:
                is_blank = True
                break
        if is_blank:
            blank_rows.append(row_values)
            blank_row_indices.append(row_num)

    return blank_rows, blank_row_indices

# Function to copy blank rows to a new workbook
def copy_blank_rows_to_new_workbook(sheet, blank_rows, new_workbook_path):
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Blank Rows"

    headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    for col, header in enumerate(headers, start=1):
        new_sheet.cell(row=1, column=col, value=header)

    for row_idx, row_data in enumerate(blank_rows, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            new_sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    new_workbook.save(new_workbook_path)
    flash(f"Blank rows copied to: {new_workbook_path}",'success')

# Function to delete blank rows from the original sheet
def delete_blank_rows_from_original(sheet, blank_row_indices):
    for row_idx in sorted(blank_row_indices, reverse=True):
        sheet.delete_rows(row_idx)

# Function to check rows where BVN No does not start with '2'
def check_bvn_not_starting_with_2(sheet):
    non_bvn_rows = []
    non_bvn_row_indices = []

    for row in sheet.iter_rows(min_row=2):
        bvn_cell_value = row[8].value  # BVN No is the 9th column (index 8)
        if isinstance(bvn_cell_value, str) and not bvn_cell_value.startswith('2'):
            non_bvn_rows.append([cell.value for cell in row])
            non_bvn_row_indices.append(row[0].row)

    return non_bvn_rows, non_bvn_row_indices

# Function to copy rows to a new worksheet
def copy_rows_to_new_workbook(sheet, rows, new_workbook_path, title):
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = title

        headers = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        for col, header in enumerate(headers, start=1):
            new_sheet.cell(row=1, column=col, value=header)

        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                new_sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        new_workbook.save(new_workbook_path)
        flash(f"Invalid Bvn copied to: {new_workbook_path}",'success')

# Function to delete rows from the original sheet
def delete_rows_from_original(sheet, row_indices):
        for row_idx in sorted(row_indices, reverse=True):
            sheet.delete_rows(row_idx)

# Function to replace gender values
def replace_gender_values(sheet, column_letter):
    for cell in sheet[column_letter]:
        if cell.value in ["M", "Male"]:
            cell.value = "001"
        elif cell.value in ["F", "Female"]:
            cell.value = "002"


# Find all Excel files in the directory
def excel_to_pipe_delimited_text(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    text_file_path = os.path.join(app.config['PROCESSED_FOLDER'], os.path.splitext(os.path.basename(file_path))[0] + '.txt')
    
    with open(text_file_path, 'w', errors='ignore') as text_file:
        for row in sheet.iter_rows(values_only=True):
            # Replace None with an empty string for each element in the row
            line = '|'.join(['' if cell is None else str(cell) for cell in row])
            text_file.write(line + '\n')

    flash(f"Conversion completed. The text file is saved as: {text_file_path}",'success')
    return text_file_path


# Individual Borrower processing
def process_individual_borrower(file_path):
    
    flash(f"Processing file: {file_path}",'success')

    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Remove duplicates
    duplicates_removed = remove_duplicates(sheet)
    if duplicates_removed:
        flash(f'Duplicates were removed.','success')
    else:
        flash(f'No duplicates found.','success')

    # Clean worksheet
    total_changes = clean_worksheet(sheet)
    print_summary(total_changes)

    # Check for and handle blank rows
    column_letters_to_check = ['A', 'C', 'D', 'F', 'I']
    blank_rows, blank_row_indices = check_blank_rows_in_columns(sheet, column_letters_to_check)
    if blank_rows:
        new_workbook_path = os.path.join(app.config['DATA_QUALITY_FOLDER'],f"{os.path.splitext(os.path.basename(file_path))[0]}_BlankRows.xlsx")
        copy_blank_rows_to_new_workbook(sheet, blank_rows, new_workbook_path)
        delete_blank_rows_from_original(sheet, blank_row_indices)

    # Check for rows where BVN No does not start with '2'
    non_bvn_rows, non_bvn_row_indices = check_bvn_not_starting_with_2(sheet)
    if non_bvn_rows:
        new_workbook_path = os.path.join(app.config['DATA_QUALITY_FOLDER'],
                                            f"{os.path.splitext(os.path.basename(file_path))[0]}_InvalidBVN.xlsx") #NonBVNStartWith2
        copy_rows_to_new_workbook(sheet, non_bvn_rows, new_workbook_path, "Non-BVN Start With 2")
        delete_rows_from_original(sheet, non_bvn_row_indices)
   
    # Format the date columns in the Excel sheet
    date_columns_to_format = ['F']
    format_date_columns(sheet, date_columns_to_format)

    # Replace gender values
    replace_gender_values(sheet, 'K')
    
    # Save the modified workbook
    workbook.save(file_path)

    # Convert to pipe-delimited text file
    excel_to_pipe_delimited_text(file_path)


# Credit Information processing
def process_credit_information(file_path):
    flash(f"Processing file: {file_path}",'success')

    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Remove duplicates
    duplicates_removed = remove_duplicates(sheet)
    if duplicates_removed:
        flash(f'Duplicates were removed.','success')
    else:
        flash(f'No duplicates found.','success')

    # Clean worksheet
    total_changes = clean_worksheet(sheet)
    print_summary(total_changes)

    # Check for and handle blank rows
    column_letters_to_check = ['A', 'B','C','D','E', 'H', 'J','M','S']
    blank_rows, blank_row_indices = check_blank_rows_in_columns(sheet, column_letters_to_check)
    if blank_rows:
        new_workbook_path = os.path.join(app.config['DATA_QUALITY_FOLDER'],f"{os.path.splitext(os.path.basename(file_path))[0]}_BlankRows.xlsx")
        copy_blank_rows_to_new_workbook(sheet, blank_rows, new_workbook_path)
        delete_blank_rows_from_original(sheet, blank_row_indices)

    # Check for rows where BVN No does not start with '2'
    non_bvn_rows, non_bvn_row_indices = check_bvn_not_starting_with_2(sheet)
    if non_bvn_rows:
        new_workbook_path = os.path.join(app.config['DATA_QUALITY_FOLDER'],
                                            f"{os.path.splitext(os.path.basename(file_path))[0]}_InvalidBVN.xlsx") #NonBVNStartWith2
        copy_rows_to_new_workbook(sheet, non_bvn_rows, new_workbook_path, "Non-BVN Start With 2")
        delete_rows_from_original(sheet, non_bvn_row_indices)
   
    # Format the date columns in the Excel sheet
    date_columns_to_format = ['D','E','P','R','U']
    format_date_columns(sheet, date_columns_to_format)
    
    # Save the modified workbook
    workbook.save(file_path)

    # Convert to pipe-delimited text file
    excel_to_pipe_delimited_text(file_path)

# Corporate Borrower processing
def process_corporate_borrower(file_path):
    flash(f"Processing file: {file_path}",'success')

    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Remove duplicates
    duplicates_removed = remove_duplicates(sheet)
    if duplicates_removed:
        flash(f'Duplicates were removed.','success')
    else:
        flash(f'No duplicates found.','success')

    # Clean worksheet
    total_changes = clean_worksheet(sheet)
    print_summary(total_changes)

    # Check for and handle blank rows
    column_letters_to_check = ['B', 'F']
    blank_rows, blank_row_indices = check_blank_rows_in_columns(sheet, column_letters_to_check)
    if blank_rows:
        new_workbook_path = os.path.join(app.config['DATA_QUALITY_FOLDER'],f"{os.path.splitext(os.path.basename(file_path))[0]}_BlankRows.xlsx")
        copy_blank_rows_to_new_workbook(sheet, blank_rows, new_workbook_path)
        delete_blank_rows_from_original(sheet, blank_row_indices)

    # Check for rows where BVN No does not start with '2'
    non_bvn_rows, non_bvn_row_indices = check_bvn_not_starting_with_2(sheet)
    if non_bvn_rows:
        new_workbook_path = os.path.join(app.config['DATA_QUALITY_FOLDER'],
                                            f"{os.path.splitext(os.path.basename(file_path))[0]}_InvalidBVN.xlsx") #NonBVNStartWith2
        copy_rows_to_new_workbook(sheet, non_bvn_rows, new_workbook_path, "Non-BVN Start With 2")
        delete_rows_from_original(sheet, non_bvn_row_indices)
   
    # Format the date columns in the Excel sheet
    date_columns_to_format = ['E']
    format_date_columns(sheet, date_columns_to_format)
    
    # Save the modified workbook
    workbook.save(file_path)


# Principal Officers processing
def process_principal_officers(file_path):
    
    flash(f"Processing file: {file_path}",'success')

    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Remove duplicates
    duplicates_removed = remove_duplicates(sheet)
    if duplicates_removed:
        flash(f'Duplicates were removed.','success')
    else:
        flash(f'No duplicates found.','success')

    # Clean worksheet
    total_changes = clean_worksheet(sheet)
    print_summary(total_changes)

    # Check for and handle blank rows
    column_letters_to_check = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V'] #Add remaining columns
    blank_rows, blank_row_indices = check_blank_rows_in_columns(sheet, column_letters_to_check)
    if blank_rows:
        new_workbook_path = os.path.join(app.config['DATA_QUALITY_FOLDER'],f"{os.path.splitext(os.path.basename(file_path))[0]}_BlankRows.xlsx")
        copy_blank_rows_to_new_workbook(sheet, blank_rows, new_workbook_path)
        delete_blank_rows_from_original(sheet, blank_row_indices)

    # Check for rows where BVN No does not start with '2'
    non_bvn_rows, non_bvn_row_indices = check_bvn_not_starting_with_2(sheet)
    if non_bvn_rows:
        new_workbook_path = os.path.join(app.config['DATA_QUALITY_FOLDER'],
                                            f"{os.path.splitext(os.path.basename(file_path))[0]}_InvalidBVN.xlsx") #NonBVNStartWith2
        copy_rows_to_new_workbook(sheet, non_bvn_rows, new_workbook_path, "Non-BVN Start With 2")
        delete_rows_from_original(sheet, non_bvn_row_indices)
   
    # Format the date columns in the Excel sheet
    date_columns_to_format = ['E','V']
    format_date_columns(sheet, date_columns_to_format)

    # Replace gender values
    replace_gender_values(sheet, ['F','W'])
    
    # Save the modified workbook
    workbook.save(file_path)

    # Convert to pipe-delimited text file
    excel_to_pipe_delimited_text(file_path)

# Guarantors Information processing
def process_guarantors_information(file_path):
    
    flash(f"Processing file: {file_path}",'success')

    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Remove duplicates
    duplicates_removed = remove_duplicates(sheet)
    if duplicates_removed:
        flash(f'Duplicates were removed.','success')
    else:
        flash(f'No duplicates found.','success')

    # Clean worksheet
    total_changes = clean_worksheet(sheet)
    print_summary(total_changes)

    # Check for and handle blank rows
    column_letters_to_check = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']
    blank_rows, blank_row_indices = check_blank_rows_in_columns(sheet, column_letters_to_check)
    if blank_rows:
        new_workbook_path = os.path.join(app.config['DATA_QUALITY_FOLDER'],f"{os.path.splitext(os.path.basename(file_path))[0]}_BlankRows.xlsx")
        copy_blank_rows_to_new_workbook(sheet, blank_rows, new_workbook_path)
        delete_blank_rows_from_original(sheet, blank_row_indices)

    # Check for rows where BVN No does not start with '2'
    non_bvn_rows, non_bvn_row_indices = check_bvn_not_starting_with_2(sheet)
    if non_bvn_rows:
        new_workbook_path = os.path.join(app.config['DATA_QUALITY_FOLDER'],
                                            f"{os.path.splitext(os.path.basename(file_path))[0]}_InvalidBVN.xlsx") #NonBVNStartWith2
        copy_rows_to_new_workbook(sheet, non_bvn_rows, new_workbook_path, "Non-BVN Start With 2")
        delete_rows_from_original(sheet, non_bvn_row_indices)
   
    # Format the date columns in the Excel sheet
    date_columns_to_format = ['I']
    format_date_columns(sheet, date_columns_to_format)

    # Replace gender values
    replace_gender_values(sheet, 'J')
    
    # Save the modified workbook
    workbook.save(file_path)

    # Convert to pipe-delimited text file
    excel_to_pipe_delimited_text(file_path)

# Upload route
@app.route('/',methods=['GET','POST'])
def index():
    return render_template('index.html')
@app.route('/index',methods=['GET','POST'])
def home():
    return render_template('home.html')
@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # Check if the POST request has the file part
        if 'file' not in request.files:
            flash('No file part', 'error')
            return redirect(request.url)
        file = request.files['file']
        
        # Check if a file is selected
        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(request.url)
        
        # Check if the file is allowed and contains a specific name
        if allowed_file(file.filename) and allowed_file_name(file.filename):
            folder = None
            if "Individual-Borrower" in file.filename:
                folder = app.config['INDIVIDUAL_BORROWER_FOLDER']
            elif "Credit-Information" in file.filename:
                folder = app.config['CREDIT_INFORMATION_FOLDER']
            elif "Corporate-Borrower" in file.filename:
                folder = app.config['CORPORATE_BORROWER_FOLDER']
            elif "Guarantors-Information" in file.filename:
                folder = app.config['GUARANTORS_INFORMATION_FOLDER']
            elif "Principal-Officers" in file.filename:
                folder = app.config['PRINCIPAL_OFFICERS_FOLDER']

            if folder:
                # Save the file to the appropriate folder
                filepath = os.path.join(folder, file.filename)
                file.save(filepath)
                start_time = time.time()
                # Process the file
                process_uploaded_file(filepath)
                # Calculate processing time
                end_time = time.time()
                processing_time = end_time - start_time
                # Flash a success message with processing time
                flash(f'File processed successfully in {processing_time:.2f} seconds', 'success')
                flash(f'{file.filename} has been processed successfully.', 'success')
                return redirect(url_for('upload_file'))

        flash('File not allowed! Upload file with correct name and file type.', 'danger')
        return redirect(request.url)
    
    return render_template('upload.html')

if __name__ == "__main__":
   app.run()
